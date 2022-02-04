import discord
import random
import openpyxl
import math
import asyncio
import mysql.connector as mysql
from PIL import Image
from uuid import uuid4
from datetime import datetime
from discord.ext import commands

client = commands.Bot(command_prefix='.', self_bot=False)
client.remove_command("help")

runestoneCost = 50
maxMobs = 10
mobValue = round(.4 * runestoneCost)
datafile = 'mobbossTest.xlsx'
coinColumn = 2
runestoneColumn = 3
mobsColumn = 4
sellFactor = 0.3
intruderChance = 10  # %
isIntruder = False
maxlevel = 99

intruder_dict = {}

class Accounts:
    coinColumn = 2
    runestoneColumn = 3
    mobsColumn = 4

    def __init__(self, name):
        workbook = openpyxl.load_workbook(filename="mobbossTest.xlsx")
        itemSheet = workbook['Items']

        self.name = name
        for row in range(2, itemSheet.max_row + 1):
            user = itemSheet['A' + str(row)].value
            if self.name == user:

                self.row = row
                break

            elif row == itemSheet.max_row:

                self.row = None
                break

        workbook.close()

    def mobs(self):
        workbook = openpyxl.load_workbook(filename="mobbossTest.xlsx")
        dungeonSheet = workbook['Dungeon']
        mobnames = []
        for column in range(2, dungeonSheet.max_column + 1):
            mobCell = dungeonSheet.cell(row=self.row, column=column).value
            if mobCell != None:
                mob = list(mobCell.split(", "))
                mobnames.append(mob[0])
        workbook.close()
        return mobnames

    def coins(self):
        workbook = openpyxl.load_workbook(filename="mobbossTest.xlsx")
        itemSheet = workbook['Items']
        workbook.close()
        return itemSheet.cell(row=self.row, column=self.coinColumn).value

    def runestones(self):
        workbook = openpyxl.load_workbook(filename="mobbossTest.xlsx")
        itemSheet = workbook['Items']
        workbook.close()
        return itemSheet.cell(row=self.row, column=self.runestoneColumn).value

    def mobcount(self):
        return len(self.mobs())


class Mob:

    def __init__(self):
        self.stats = []
        self.growth = []

    def level_up(self, levelAdd=1):
        growth = []
        workbook = openpyxl.load_workbook(filename="mobbossTest.xlsx")
        statSheet = workbook['MobStats']
        growthSheet = workbook['GrowthRates']
        mobSheet = workbook['Mobs']
        for column in range(2, statSheet.max_column):
            growth.append(growthSheet.cell(row=self.row, column=column).value)
        for lvl in range(levelAdd):
            for i in range(len(self.stats)):
                roll = random.randint(0, 100)
                if growth[i] >= roll:
                    if i == 4:
                        self.stats[i] += 5
                    if i != 4:
                        self.stats[i] += 1
                        if self.stats[i] >= 99:
                            self.stats[i] = 99

        self.level = statSheet.cell(row=self.row, column=7).value + levelAdd
        if self.info[6] >= 100:
            self.info[6] -= 100
        self.info[5] += 1

        workbook.save(filename=datafile)
        pass

    def create_mob(self, user):
        workbook = openpyxl.load_workbook(filename="mobbossTest.xlsx")
        rollSheet = workbook['MobChance']
        statSheet = workbook['MobStats']
        growthSheet = workbook['GrowthRates']
        self.stats = []
        self.growth = []
        roll = random.uniform(0.0, 100.0)
        # rolling mobs
        for row in range(2, rollSheet.max_row):
            if rollSheet.cell(row=row + 1, column=7).value is not None:
                if roll < rollSheet.cell(row=row, column=7).value and roll > rollSheet.cell(row=row + 1,
                                                                                            column=7).value:
                    mobname = rollSheet["A" + str(row)].value

        # finding mob on statsheet
        for row in range(2, statSheet.max_row + 1):
            mob = statSheet['A' + str(row)].value
            if mobname == mob:
                self.row = row

        # assigning statsheet to mob
        for column in range(2, statSheet.max_column):
            self.stats.append(statSheet.cell(row=self.row, column=column).value)
            self.growth.append(growthSheet.cell(row=self.row, column=column).value)

        self.level = 1
        xp = 0
        rarity = rollSheet.cell(row=self.row, column=2).value
        color = rollSheet.cell(row=self.row, column=3).value
        # creating info
        self.info = [uuid4().hex,
                     user.name,
                     mobname,
                     rarity,
                     color,
                     self.level,
                     xp]

        # Adding Levels
        levelAdd = random.randint(0, 5)
        self.level_up(levelAdd)
        self.info[5] = self.level

        workbook.close()

    # gets mob from existing
    def get_mob(self, number, user):
        workbook = openpyxl.load_workbook(filename="mobbossTest.xlsx")
        dungeonSheet = workbook['Dungeon']
        mobSheet = workbook['Mobs']

        mob = dungeonSheet.cell(row=user.row, column=number + 1).value
        mob = list(mob.split(", "))
        mobID = mob[1]

        for row in range(2, mobSheet.max_row + 1):
            if mobSheet.cell(row=row, column=1).value == mobID:
                # MAYBE MAKE DICTIONARIES WITH TITLES, PROBALBY BETTER.
                self.stats = [mobSheet.cell(row=row, column=8).value,
                              mobSheet.cell(row=row, column=9).value,
                              mobSheet.cell(row=row, column=10).value,
                              mobSheet.cell(row=row, column=11).value,
                              mobSheet.cell(row=row, column=12).value]
                self.info = [mobSheet.cell(row=row, column=1).value,  # ID
                             mobSheet.cell(row=row, column=2).value,  # Owner
                             mobSheet.cell(row=row, column=3).value,  # Name
                             mobSheet.cell(row=row, column=4).value,  # Rarity
                             mobSheet.cell(row=row, column=5).value,  # Color
                             mobSheet.cell(row=row, column=6).value,  # LVL
                             mobSheet.cell(row=row, column=7).value]  # XP
                self.row = row
                break
        workbook.close()

    def get_mob_ID(self, mobID):
        workbook = openpyxl.load_workbook(filename="mobbossTest.xlsx")
        mobSheet = workbook['Mobs']
        for row in range(2, mobSheet.max_row + 1):
            if mobSheet.cell(row=row, column=1).value == mobID:
                # MAYBE MAKE DICTIONARIES WITH TITLES, PROBALBY BETTER.
                self.stats = [mobSheet.cell(row=row, column=8).value,
                              mobSheet.cell(row=row, column=9).value,
                              mobSheet.cell(row=row, column=10).value,
                              mobSheet.cell(row=row, column=11).value,
                              mobSheet.cell(row=row, column=12).value]
                self.info = [mobSheet.cell(row=row, column=1).value,  # ID
                             mobSheet.cell(row=row, column=2).value,  # Owner
                             mobSheet.cell(row=row, column=3).value,  # Name
                             mobSheet.cell(row=row, column=4).value,  # Rarity
                             mobSheet.cell(row=row, column=5).value,  # Color
                             mobSheet.cell(row=row, column=6).value,  # LVL
                             mobSheet.cell(row=row, column=7).value]  # XP
                self.row = row
                break
        workbook.close()


class Intruder:
    def __init__(self):
        self.info = []
        self.stats = []
        self.growth = []
        workbook = openpyxl.load_workbook(filename="mobbossTest.xlsx")
        intruderSheet = workbook["Intruders"]
        roll = random.randint(2, intruderSheet.max_row)
        for column in range(1, intruderSheet.max_column + 1):

            if 1 <= column <= 7:
                self.info.append(intruderSheet.cell(row=roll, column=column).value)
            if 8 <= column <= 12:
                self.stats.append(intruderSheet.cell(row=roll, column=column).value)
            if 13 <= column <= 17:
                self.growth.append(intruderSheet.cell(row=roll, column=column).value)
        level = self.info[5]
        toMaxLevel = maxlevel - level
        lvlAdd = random.randint(0, 10)
        for j in range(lvlAdd):
            for i in range(len(self.stats)):
                roll = random.randint(0, 100)
                if self.growth[i] >= roll:
                    if i == 4:
                        self.stats[i] += 5
                    if i != 4:
                        self.stats[i] += 1
                        if self.stats[i] >= 99:
                            self.stats[i] = 99
        self.info[5] += lvlAdd
        workbook.close()


async def intruderAlert(message):
    guild_id = message.guild.id
    global intruder_dict

    if intruder_dict[guild_id][0] != True:
        intruder = Intruder()

        intruderName = intruder.info[2]
        intruderLevel = intruder.info[5]

        alert = 0xeb321e
        alertFile = discord.File("./Images/duel.png", filename="image.png")
        intruderFile = discord.File("./Images/%s.png" % intruderName, filename="image2.png")
        authorName = "Intruder Alert!"
        embed = discord.Embed(title="A `%s` has entered your dungeon!" % intruderName, description="Repel them!",
                              color=alert)
        embed.set_author(name=authorName, icon_url="attachment://image.png")
        embed.set_thumbnail(url="attachment://image2.png")
        embed.add_field(name="⭐LVL:", value="`%s`" % intruderLevel, inline=False)

        embed.set_footer(text="Use .battle <mob number> to repel the intruder")

        newMessage = await message.channel.send(embed=embed, files=[alertFile, intruderFile])
        intruder_dict[guild_id] = [True, intruder, newMessage]
        print(intruder_dict)
        await asyncio.sleep(30)
        if intruder_dict[guild_id][0] == True:
            await newMessage.delete()
            intruder_dict[guild_id][0] = False
            del intruder


# BATTLE FUNCTIONS--------------------------------------------------------------------------------


async def doBattle(fighter1, fighter2):
    SPDlimit = 500
    HP1, HP2 = fighter1.stats[4], fighter2.stats[4]
    ATK1, ATK2 = fighter1.stats[0], fighter2.stats[0]
    DEF1, DEF2 = fighter1.stats[1], fighter2.stats[1]
    SPD1, SPD2 = fighter1.stats[2], fighter2.stats[2]
    ACC1, ACC2 = fighter1.stats[3], fighter2.stats[3]

    HP = [[HP1, HP2]]
    DMG = [[0, 0]]
    hitType = [['', '']]
    winner = None
    SPDcount1 = 0
    SPDcount2 = 0
    rounds = 0
    while winner is None:
        rounds += 1

        # SPEED
        SPDcount1 += SPD1
        SPDcount2 += SPD2
        if SPDcount1 > SPDlimit:
            SPDcount1 -= SPDlimit
        if SPDcount2 > SPDlimit:
            SPDcount2 -= SPDlimit
        # TURNS
        if SPDcount2 == SPDcount1:
            roll = random.randint(1, 2)
            if roll == 1:
                winner, HP1, HP2, hitType1, hitType2 = await doRound(fighter1, HP1, ATK1, DEF1, SPD1, ACC1, fighter2,
                                                                     HP2, ATK2, DEF2, SPD2, ACC2)
            if roll == 2:
                winner, HP2, HP1, hitType2, hitType1 = await doRound(fighter2, HP2, ATK2, DEF2, SPD2, ACC2, fighter1,
                                                                     HP1, ATK1, DEF1, SPD1, ACC1)

        elif SPDcount1 > SPDcount2:
            winner, HP1, HP2, hitType1, hitType2 = await doRound(fighter1, HP1, ATK1, DEF1, SPD1, ACC1, fighter2, HP2,
                                                                 ATK2, DEF2, SPD2, ACC2)
        elif SPDcount2 > SPDcount1:
            winner, HP2, HP1, hitType2, hitType1 = await doRound(fighter2, HP2, ATK2, DEF2, SPD2, ACC2, fighter1, HP1,
                                                                 ATK1, DEF1, SPD1, ACC1)

        hitType.append([hitType1, hitType2])
        HP.append([HP1, HP2])
        previous = rounds - 1
        DMGto1 = HP[previous][0] - HP[rounds][0]
        DMGto2 = HP[previous][1] - HP[rounds][1]
        DMG.append([DMGto1, DMGto2])

    if winner == fighter1:
        loser = fighter2
    elif winner == fighter2:
        loser = fighter1

    xp = await calculate_xp(winner.info[5], loser.info[5])

    return winner, HP, DMG, hitType, xp


async def doDamage(ATK1, ACC1, DEF2, HP2, SPD2):
    c = 1
    modifier = 1
    hitType = "Hit!"
    critChance = ACC1 / SPD2
    roll = random.randint(0, 100)
    if roll <= critChance:
        modifier = 1.5
        hitType = "Critical Hit!"
    DMG = round((((ATK1 ** 2) * (1 - (DEF2 / (c + DEF2)))) * modifier) / 2)
    HP2 -= DMG
    return HP2, hitType


async def isHit(ACC):
    roll = random.randint(0, 100)
    if roll <= ACC:
        return True
    else:
        return False


async def doRound(fighter1, HP1, ATK1, DEF1, SPD1, ACC1, fighter2, HP2, ATK2, DEF2, SPD2, ACC2):
    winner = None

    # FIGHTER1 GOES FIRST
    if await isHit(ACC1):

        HP2, hitType2 = await doDamage(ATK1, ACC1, DEF2, HP2, SPD2)
        if HP2 <= 0:
            winner = fighter1
            hitType1 = "None"
            return winner, HP1, HP2, hitType1, hitType2
    else:
        hitType2 = "Miss!"
    if await isHit(ACC2):
        HP1, hitType1 = await doDamage(ATK2, ACC2, DEF1, HP1, SPD1)
        if HP1 <= 0:
            winner = fighter2
            return winner, HP1, HP2, hitType1, hitType2
    else:
        hitType1 = "Miss!"

    return winner, HP1, HP2, hitType1, hitType2


async def calculate_xp(winnerLVL, loserLVL):
    difficulty = loserLVL - winnerLVL

    if difficulty >= -3 or difficulty <= 3:
        xp = 25
    elif 3 < difficulty <= 6:
        xp = 30
    elif difficulty > 6:
        xp = 35
    elif -3 > difficulty >= -6:
        xp = 20
    elif difficulty < -6:
        xp = 10
    return xp


# END BATTLE FUNCTIONS---------------------------------------------------------------------
async def train_xp(mobID, xp, message):
    print('line395 test, train xp called')
    print(xp)
    workbook = openpyxl.load_workbook(filename=datafile, data_only=True)
    mobSheet = workbook['Mobs']
    mobInst = Mob()
    mobInst.get_mob_ID(mobID)
    print(mobID)

    for row in range(2, mobSheet.max_row + 1):
        if mobSheet.cell(row=row, column=1).value == mobID:
            mobRow = row
            break
    print('xp:%s' % mobInst.info[6])
    mobSheet.cell(row=mobRow, column=7).value += xp
    mobInst.info[6] += xp
    print('new_xp:%s' % mobInst.info[6])
    print('new_xp_check:%s' % mobSheet.cell(row=mobRow, column=7).value)
    if mobSheet.cell(row=mobRow, column=7).value >= 100:
        oldStats = mobInst.stats[:]
        oldInfo = mobInst.info[:]
        mobInst.level_up(1)

        for column in range(8, 13):
            mobSheet.cell(row=mobInst.row, column=column).value = mobInst.stats[column - 8]
        mobSheet.cell(row=mobInst.row, column=6).value = mobInst.info[5]
        mobSheet.cell(row=mobInst.row, column=7).value = mobInst.info[6]

        newStats = mobInst.stats
        newInfo = mobInst.info

        await print_level_up(message, oldStats, oldInfo, newStats, newInfo)
    workbook.save(filename=datafile)
    print('fuck')



async def print_level_up(message, oldStats, oldInfo, newStats, newInfo):
    user = Accounts(message.author.name)
    change = []
    for i in range(len(newStats)):
        if newStats[i] - oldStats[i] != 0:
            change.append("🌟")
        else:
            change.append("")
    oldLevel = oldInfo[5]
    newLevel = newInfo[5]
    rarity = newInfo[3]
    color = newInfo[4]

    sixteenIntegerHex = int(color.replace("#", ""), 16)
    readableHex = int(hex(sixteenIntegerHex), 0)

    File = discord.File("./Images/%s.png" % newInfo[2], filename="image.png")
    rarityFile = discord.File("./Images/%s.png" % rarity, filename="rarity.png")

    embed = discord.Embed(title="🌟Level: `%s -> %s` " % (oldLevel, newLevel), color=readableHex)
    embed.set_author(name="%s's %s Leveled Up!" % (newInfo[1], newInfo[2]), icon_url="attachment://rarity.png")
    embed.set_thumbnail(url="attachment://image.png")
    embed.add_field(name="❤HP:", value="%s`%s` -> `%s`" % (change[4], oldStats[4], newStats[4]), inline=False)
    embed.add_field(name="⚔ATK:", value="%s`%s` -> `%s`" % (change[0], oldStats[0], newStats[0]), inline=False)
    embed.add_field(name="🛡DEF:", value="%s`%s` -> `%s`" % (change[1], oldStats[1], newStats[1]), inline=False)
    embed.add_field(name="🪶SPD:", value="%s`%s` -> `%s`" % (change[2], oldStats[2], newStats[2]), inline=False)
    embed.add_field(name="🎯ACC:", value="%s`%s` -> `%s`" % (change[3], oldStats[3], newStats[3]), inline=False)
    await message.channel.send(embed=embed, files=[File, rarityFile])


@client.event
async def on_ready():
    await client.change_presence(activity=discord.Game(name=".mobboss"))
    for guild in client.guilds:
        #[Intruder state, intruder class, message.id]
        intruder_dict[guild.id] = [False, None, 0]
    print('Bot is ready.')
    return


# MESSAGE POINTS----REWRITTEN-------------------
@client.event
async def on_message(message):
    user = Accounts(message.author.name)
    context = await client.get_context(message)
    # POINTS
    if message.author.name != client.user.name and user.row is not None and not context.valid:

        workbook = openpyxl.load_workbook(filename=datafile, data_only=True)
        itemsSheet = workbook['Items']
        trainingSheet = workbook['TrainingCache']
        mobID = trainingSheet.cell(row=user.row, column=2).value

        length = len(message.content)
        L, a, k, x0 = 10, 13.4, 0.024, 10
        mu = L / (1 + (a * math.exp(-k * (length - x0))))
        coinsAdd = round(random.normalvariate(mu, 1))
        xp = coinsAdd
        if coinsAdd < 0:
            coinsAdd = 0

        itemsSheet.cell(row=user.row, column=coinColumn).value += coinsAdd

        # Training
        #if trainingSheet.cell(row=user.row, column=2).value is not None:
            #print('line490 test - trainsheet is not none')
            #await train_xp(mobID, xp, message)

        # INTRUDER ALERTS
        print('intruder')
        roll = random.randint(0, 100)
        print(roll)
        print(intruderChance)
        if roll <= intruderChance:
            print('True')
            await intruderAlert(message)
        workbook.save(filename=datafile)

    del user
    await client.process_commands(message)


# ---------------------------------------------------

# VOICE POINTS----------checked-----------------
@client.event
async def on_voice_state_update(member, before, after):
    workbook = openpyxl.load_workbook(filename=datafile, data_only=True)
    voiceSheet = workbook['VoiceCache']
    itemSheet = workbook['Items']
    user = Accounts(member.name)
    timeMultiplier, timeMinimum, timeCap = 8, 1200, 7200
    row = user.row

    if before.channel is None and after.channel is not None:
        voiceSheet.cell(row=row, column=2).value = datetime.now()

    elif before.channel is not None and after.channel is None:
        change = datetime.now() - voiceSheet.cell(row=row, column=2).value
        timeSpent = change.seconds

        if timeSpent > timeCap:
            timeSpent = timeCap
        coinsAdd = round(timeMultiplier * timeSpent / timeMinimum)
        if timeSpent < timeMinimum:
            points = 0

        itemSheet.cell(row=row, column=2).value += coinsAdd
    del user
    workbook.save(filename=datafile)


# ----------------------------------------------------
@client.group(invoke_without_command=True)
async def help(ctx):
    File = discord.File("./Images/Helping Hand.png", filename="image.png")
    embed = discord.Embed(title="Help", description="Use `.help <category>` for extended information.", colour=0x114a1d)
    embed.set_author(name="Helping Hand", icon_url="attachment://image.png")
    embed.set_thumbnail(url="attachment://image.png")
    embed.add_field(name="Account", value="`.mobboss`", inline=False)
    embed.add_field(name="Mobs", value="`.mobs` `.battle` `.train`", inline=False)
    embed.add_field(name="Market", value="`.buy` `.sell`", inline=False)
    embed.add_field(name="Runestones", value="`open`", inline=False)
    await ctx.send(embed=embed, file=File)


@help.command(aliases=['account'])
async def Account(ctx):
    embed = discord.Embed(title="Account", description="Gives account information", colour=0x114a1d)
    embed.add_field(name=".mobboss", value="Creates account or gives information about your account", inline=False)
    await ctx.send(embed=embed)


@help.command(aliases=['mobs'])
async def Mobs(ctx):
    embed = discord.Embed(title="Mobs", description="Interact with your Mobs", colour=0x114a1d)
    embed.add_field(name=".mobs", value="Look at a list of all mobs in your dungeon", inline=False)
    embed.add_field(name=".mobs <number>", value="Look at a specific mob chosen from your list", inline=False)
    embed.add_field(name=".train <number>",
                    value="Set up a training regiment for your mob, allowing it to gain xp passively")
    embed.add_field(name=".battle <number>", value="Battle an intruder with chosen mob", inline=False)
    await ctx.send(embed=embed)


@help.command(aliases=['market'])
async def Market(ctx):
    embed = discord.Embed(title="Market", description="Buy and Sell items", colour=0x114a1d)
    embed.add_field(name=".buy <item> <amount>", value="Buys <amount> of <items>", inline=False)
    embed.add_field(name=".sell <item> <amount>", value="Sells <amount> of <items>", inline=False)
    embed.add_field(name=".sell mobs <number>", value="Sell a specific mob chosen from your list", inline=False)
    await ctx.send(embed=embed)


@help.command(aliases=['runestones'])
async def Runestones(ctx):
    embed = discord.Embed(title="Runestones", description="Runestones!", colour=0x114a1d)
    embed.add_field(name=".open <runestonetype>", value="Summon a mob from a runestone in your inventory", inline=False)
    await ctx.send(embed=embed)


# Creates new 'account'
@client.command(brief='Creates a Mob Boss account',
                description="Use this command to create an account with The Mob Boss")
async def mobboss(ctx, show='no'):
    if show == 'show':
        destination = ctx
    else:
        destination = ctx.author

    user = Accounts(ctx.author.name)
    workbook = openpyxl.load_workbook(filename=datafile, data_only=True)
    dungeonSheet = workbook['Dungeon']
    itemSheet = workbook['Items']
    voiceSheet = workbook['VoiceCache']
    trainingSheet = workbook['TrainingCache']
    hasAccount = False

    File = discord.File("./Images/Helping Hand.png", filename="image.png")

    # find user
    if user.row is not None:
        hasAccount = True

    if hasAccount == True:
        title = "His Holiness, Lich **%s**" % ctx.author.name
        description = "Always happy to lend a hand."
        green = 0x114a1d
        authorName = "Helping Hand"
        embed = discord.Embed(title=title, description=description, color=green)
        embed.set_author(name=authorName, icon_url="attachment://image.png")
        embed.set_thumbnail(url="attachment://image.png")
        embed.add_field(name="Mobs", value="**%s** mobs currently guard in your dungeon" % user.mobcount(),
                        inline=False)
        embed.add_field(name="Runestones", value="You currently have **%s** unopened Runestones" % user.runestones(),
                        inline=False)
        embed.add_field(name="Coins", value="**%s** coins lay in your coffers" % user.coins(), inline=False)
        embed.set_footer(text="Use .help for a list of commands")
        await destination.send(embed=embed, file=File)

    if hasAccount == False:
        row = dungeonSheet.max_row + 1
        dungeonSheet.cell(row=row, column=1).value = ctx.author.name
        itemSheet.cell(row=row, column=1).value = ctx.author.name
        voiceSheet.cell(row=row, column=1).value = ctx.author.name
        trainingSheet.cell(row=row, column=1).value = ctx.author.name
        itemSheet.cell(row=row, column=2).value = 0
        itemSheet.cell(row=row, column=3).value = 1
        itemSheet.cell(row=row, column=4).value = 0

        embed = discord.Embed(title="Your Unholiness!",
                              description="It's every dismembered hand's dream to become a Lich's right hand... hand. I'll get you started with the basics of running your own **Dungeon**",
                              color=0x114a1d)
        embed.set_author(name="Helping Hand",
                         icon_url="attachment://image.png")
        embed.set_thumbnail(
            url="attachment://image.png")
        embed.add_field(name="Coins",
                        value="`Coins` are periodically awarded to you for participating in Text or Voice Channels...whatever those are.   `Coins` can be used to purchase `Runestones` and train `Mobs`",
                        inline=True)
        embed.add_field(name="Runestones",
                        value="`Runestones` contain the essence of `Mobs`. You can open `Runestones` with your Lich magics to summon new `Mobs`",
                        inline=False)
        embed.add_field(name="Mobs",
                        value="What good is a dungeon if it doesn't have `Mobs` defending it...from the ridicule of your fellow Liches? Collect rare and powerful `Mobs` to win bragging rights.",
                        inline=False)
        embed.add_field(name="Getting Started",
                        value="Here's a Runestone from my previous employer. Use `.open runestone` now to summon your first `Mob`",
                        inline=False)
        embed.set_footer(text="Use .help for a list of commands")
        await destination.send(embed=embed, file=File)
    del user
    workbook.save(filename=datafile)


# Looking at mobs.
@client.command()
async def mobs(ctx, choice=0, show='no'):
    if show == 'show':
        destination = ctx
    else:
        destination = ctx.author

    user = Accounts(ctx.author.name)
    mobInst = Mob()
    row = user.row
    msg = ''
    if choice != 0:
        mob = user.mobs()[choice - 1]
        mobInst.get_mob(choice, user)
        level = mobInst.info[5]
        rarity = mobInst.info[3]
        color = mobInst.info[4]

        sixteenIntegerHex = int(color.replace("#", ""), 16)
        readableHex = int(hex(sixteenIntegerHex), 0)

        File = discord.File("./Images/%s.png" % mobInst.info[2], filename="image.png")
        rarityFile = discord.File("./Images/%s.png" % rarity, filename="rarity.png")

        embed = discord.Embed(title="Level: `%s`     Experience: `%s` " % (level, mobInst.info[6]), color=readableHex)
        embed.set_author(name="%s's %s" % (user.name, mobInst.info[2]), icon_url="attachment://rarity.png")
        embed.set_thumbnail(url="attachment://image.png")
        embed.add_field(name="❤HP:", value="`%s`" % mobInst.stats[4], inline=False)
        embed.add_field(name="⚔ATK:", value="`%s/99`" % mobInst.stats[0], inline=False)
        embed.add_field(name="🛡DEF:", value="`%s/99`" % mobInst.stats[1], inline=False)
        embed.add_field(name="🪶SPD:", value="`%s/99`" % mobInst.stats[2], inline=False)
        embed.add_field(name="🎯ACC:", value="`%s/99`" % mobInst.stats[3], inline=False)
        await destination.send(embed=embed, files=[File, rarityFile])

    else:
        embed = discord.Embed(title=' ', colour=0x78456f)
        embed.set_author(name="%s's Dungeon" % user.name)
        msg = ''
        for i in range(user.mobcount()):
            mobInst.get_mob(i + 1, user)
            msg += "%s. **%s** - Lvl %s\n" % (i + 1, user.mobs()[i], mobInst.info[5])

        embed.add_field(name="Mobs", value=msg, inline=True)
        await destination.send(embed=embed)
    del user
    del mobInst


@client.command()
async def sell(ctx, item, amount=1, show='no'):
    if show == 'show':
        destination = ctx
    else:
        destination = ctx.author

    user = Accounts(ctx.author.name)
    workbook = openpyxl.load_workbook(filename=datafile, data_only=True)
    itemSheet = workbook['Items']
    dungeonSheet = workbook['Dungeon']
    mobSheet = workbook['Mobs']
    row = user.row

    file = discord.File("./Images/Coin Purse.png", filename="image.png")
    amount = int(amount)
    if amount >= 0:
        if item == 'runestone':
            if amount <= user.runestones():
                sellValue = amount * runestoneCost
                itemSheet.cell(row=row, column=coinColumn).value += round((amount * runestoneCost * sellFactor))
                itemSheet.cell(row=row, column=runestoneColumn).value -= amount

                embed = discord.Embed(title="You're the best business partner a goblin could ask for!",
                                      description="Here's a summary of our transaction:", color=0xd6ba00)
                embed.set_author(name="Master of Business Acumen", icon_url="attachment://image.png")
                embed.set_thumbnail(url="attachment://image.png")
                embed.add_field(name="Sold", value="%s runestones" % amount, inline=True)
                embed.add_field(name="Total Value", value=str(sellValue * sellFactor), inline=True)
                embed.set_footer(text="Coins: %s\nRunestones: %s" % (
                itemSheet.cell(row=row, column=coinColumn).value, itemSheet.cell(row=row, column=runestoneColumn).value))
                await destination.send(embed=embed, files=[file])

            else:
                embed = discord.Embed(title="That's not possible, hoss.", description="You haven't enough runestones.",
                                      color=0xd6ba00)
                embed.set_author(name="Master of Business Acumen", icon_url="attachment://image.png")
                embed.set_thumbnail(url="attachment://image.png")
                embed.set_footer(text="Coins: %s\nRunestones: %s" % (
                itemSheet.cell(row=row, column=coinColumn).value, itemSheet.cell(row=row, column=runestoneColumn).value))
                await destination.send(embed=embed, files=[file])

        elif item == 'mobs':
            mobInst = Mob()
            mobInst.get_mob(amount, user)
            mobname = user.mobs()[amount - 1]

            itemSheet.cell(row=row, column=2).value += round(mobValue * sellFactor)
            for column in range(amount + 1, maxMobs + 2):
                dungeonSheet.cell(row=row, column=column).value = dungeonSheet.cell(row=row, column=column + 1).value

            mobSheet.move_range("A%s:Z%s" % (mobInst.row + 1, mobSheet.max_row + 1), rows=-1)

            embed = discord.Embed(title="You're the best business partner a goblin could ask for!",
                                  description="Here's a summary of our transaction:", color=0xd6ba00)
            embed.set_author(name="Master of Business Acumen", icon_url="attachment://image.png")
            embed.set_thumbnail(url="attachment://image.png")
            embed.add_field(name="Sold", value=mobname, inline=True)
            embed.add_field(name="Total Value", value=str(mobValue * sellFactor), inline=True)
            embed.set_footer(text="Coins: %s\nRunestones: %s" % (
            itemSheet.cell(row=row, column=coinColumn).value, itemSheet.cell(row=row, column=runestoneColumn).value))
            await destination.send(embed=embed, files=[file])
            del mobInst
        else:
            embed = discord.Embed(title="There's not much of a market for that.",
                                  description="Here's a list of things I'll buy from you and for how much:", color=0xd6ba00)
            embed.set_author(name="Master of Business Acumen", icon_url="attachment://image.png")
            embed.set_thumbnail(url="attachment://image.png")
            embed.add_field(name="Item", value="`Runestone`", inline=True)
            embed.add_field(name="Price", value=str(runestoneCost), inline=True)
            embed.set_footer(text="Coins: %s\nRunestones: %s" % (
            itemSheet.cell(row=row, column=coinColumn).value, itemSheet.cell(row=row, column=runestoneColumn).value))
            await destination.send(embed=embed, files=[file])
    # if item == mob
    del user
    workbook.save(filename=datafile)


@client.command()
async def buy(ctx, item, amount=1, show='no'):
    if show == 'show':
        destination = ctx
    else:
        destination = ctx.author
    user = Accounts(ctx.author.name)
    workbook = openpyxl.load_workbook(filename=datafile, data_only=True)
    itemSheet = workbook['Items']
    row = user.row
    file = discord.File("./Images/Coin Purse.png", filename="image.png")
    amount = int(amount)
    if amount >= 0:
        if item == 'runestone':
            if amount * runestoneCost <= itemSheet.cell(row=row, column=2).value:
                itemSheet.cell(row=row, column=2).value -= amount * runestoneCost
                itemSheet.cell(row=row, column=3).value += amount
                embed = discord.Embed(title="I'll be rich in no time!", description="Here's a summary of our transaction:",
                                      color=0xd6ba00)
                embed.set_author(name="Master of Business Acumen", icon_url="attachment://image.png")
                embed.set_thumbnail(url="attachment://image.png")
                embed.add_field(name="Purchased", value=str(amount) + " runestones", inline=True)
                embed.add_field(name="Total Cost", value=str(amount * runestoneCost), inline=True)
                embed.set_footer(text="Coins: %s\nRunestones: %s" % (
                itemSheet.cell(row=row, column=coinColumn).value, itemSheet.cell(row=row, column=runestoneColumn).value))
                await destination.send(embed=embed, files=[file])
            else:
                embed = discord.Embed(title="No can do, partner!", description="You haven't enough coin.", color=0xd6ba00)
                embed.set_author(name="Master of Business Acumen", icon_url="attachment://image.png")
                embed.set_thumbnail(url="attachment://image.png")
                embed.set_footer(text="Coins: %s\nRunestones: %s" % (
                itemSheet.cell(row=row, column=coinColumn).value, itemSheet.cell(row=row, column=runestoneColumn).value))
                await destination.send(embed=embed, files=[file])

        else:
            embed = discord.Embed(title="At the moment my inventory is pretty limited. ",
                                  description="Here's a list of things you can buy and for how much", color=0xd6ba00)
            embed.set_author(name="Master of Business Acumen", icon_url="attachment://image.png")
            embed.set_thumbnail(url="attachment://image.png")
            embed.add_field(name="Inventory", value="`Runestone`", inline=True)
            embed.add_field(name="Price", value=str(runestoneCost), inline=True)
            embed.set_footer(text="Coins: %s\nRunestones: %s" % (
            itemSheet.cell(row=row, column=coinColumn).value, itemSheet.cell(row=row, column=runestoneColumn).value))
            await destination.send(embed=embed, files=[file])

    workbook.save(filename=datafile)


@client.command()
async def open(ctx, item):
    user = Accounts(ctx.author.name)
    workbook = openpyxl.load_workbook(filename=datafile, data_only=True)
    itemSheet = workbook['Items']
    dungeonSheet = workbook['Dungeon']
    mobSheet = workbook['Mobs']
    row = user.row
    mobInst = Mob()
    numRs = itemSheet.cell(row=row, column=3).value
    if item == "runestone":
        if numRs > 0:
            itemSheet.cell(row=row, column=3).value -= 1
            maxColNum = 2
            for colNum in range(2, dungeonSheet.max_column + 1):
                if dungeonSheet.cell(row=row, column=colNum).value != None:
                    maxColNum += 1
                else:
                    break
            if maxColNum <= 11:

                mobInst.create_mob(user)
                rowToWrite = mobSheet.max_row + 1
                for column in range(1, 8):
                    mobSheet.cell(row=rowToWrite, column=column).value = mobInst.info[column - 1]
                for column in range(8, 13):
                    mobSheet.cell(row=rowToWrite, column=column).value = mobInst.stats[column - 8]
                mobname = mobInst.info[2]
                mobID = mobInst.info[0]
                dungeonSheet.cell(row=row, column=maxColNum).value = "%s, %s" % (mobname, mobID)
                File = discord.File("./Images/%s.png" % mobname, filename="image.png")
                runeFile = discord.File("./Images/Rune Stone.png", filename="runestone.png")
                embed = discord.Embed(title="Congratulations!",
                                      description="%s has just summoned a %s!" % (ctx.author.name, mobname),
                                      color=0xaa5527)
                embed.set_author(name=mobname + " Summoned!", icon_url="attachment://runestone.png")
                embed.set_thumbnail(url="attachment://image.png")
                embed.set_footer(text="Coins: %s\nRunestones: %s" % (itemSheet.cell(row=row, column=coinColumn).value,
                                                                     itemSheet.cell(row=row,
                                                                                    column=runestoneColumn).value))
                await ctx.send(embed=embed, files=[runeFile, File])

            else:

                File = discord.File("./Images/Helping Hand.png", filename="image.png")
                embed = discord.Embed(title="His Holiness, Lich **%s**" % ctx.author.name,
                                      description="There appears to be a problem.", color=0x114a1d)
                embed.set_author(name="Helping Hand", icon_url="attachment://image.png")
                embed.set_thumbnail(url="attachment://image.png")
                embed.add_field(name="Dungeons Filled",
                                value="There are too many `mobs` residing in your dungeon. You'll have to sell a mob with `.sell mobs <mob number>` before you can open another runestone. ",
                                inline=False)
                embed.set_footer(text="Use .help for a list of commands")
                await ctx.send(embed=embed, file=File)


        else:
            await ctx.send(
                "Sorry, %s! You do not currently have any runestones. Purchase runestones with points`.buy runestone <amount>` to get more runestones" % ctx.author.name)

    workbook.save(filename=datafile)


@client.command()
async def train(ctx, choice=1):
    workbook = openpyxl.load_workbook(filename=datafile, data_only=True)
    trainingSheet = workbook['TrainingCache']
    user = Accounts(ctx.author.name)
    row = user.row
    mobInst = Mob()
    mobInst.get_mob(choice, user)

    trainingSheet.cell(row=row, column=2).value = mobInst.info[0]

    File = discord.File("./Images/Helping Hand.png", filename="image.png")
    mobFile = discord.File("./Images/%s.png" % mobInst.info[2], filename="image2.png")
    green = 0x114a1d
    embed = discord.Embed(title="**%s** is in good hands, Lich **%s**" % (mobInst.info[2], ctx.author.name),
                          description="Your %s [lvl %s] is currently being trained!" % (
                          mobInst.info[2], mobInst.info[5]), color=green)
    embed.set_author(name="Helping Hand", icon_url="attachment://image.png")
    embed.set_thumbnail(url="attachment://image2.png")
    await ctx.author.send(embed=embed, files=[File, mobFile])

    workbook.save(filename=datafile)


@client.command()
async def battle(ctx, choice: int):
    global intruder_dict
    await ctx.message.delete()
    workbook = openpyxl.load_workbook(filename=datafile, data_only=True)
    itemSheet = workbook['Items']
    mobSheet = workbook['Mobs']
    user = Accounts(ctx.author.name)
    if intruder_dict[ctx.guild.id][0] == True:
        mobInst = Mob()
        mobInst.get_mob(choice, user)

        #BATTLE
        winner, HP, DMG, hitType, xp = await doBattle(mobInst, intruder_dict[ctx.guild.id][1])
        msg_to_delete = intruder_dict[ctx.guild.id][2]
        await msg_to_delete.delete()
        rounds = len(HP)

        # making arena image
        mobname1 = mobInst.info[2]
        mobname2 = intruder_dict[ctx.guild.id][1].info[2]
        arena = "Arena2"
        arenaImage = Image.open('./Images/%s.png' % arena)
        mob1 = Image.open('./Images/%s.png' % mobname1)
        mob2 = Image.open('./Images/%s.png' % mobname2)

        background_im = arenaImage.copy()
        background_im.paste(mob1, (0, 15), mask=mob1)
        background_im.paste(mob2, (190, 15), mask=mob2)

        background_im.save('./Images/newArena.png')

        File = discord.File("./Images/duel.png", filename="image.png")
        winnerFile = discord.File("./Images/%s.png" % winner.info[2], filename="winnerimage.png")
        arenaFile = discord.File("./Images/newArena.png", filename="arenaimage.png")
        embed = discord.Embed(title="%s's `%s` has won!" % (winner.info[1], winner.info[2]), description="Battle Log",
                              color=0xeb321e)
        embed.set_author(name="Intruder Alert", icon_url="attachment://image.png")
        embed.set_thumbnail(url="attachment://winnerimage.png")
        embed.set_image(url="attachment://arenaimage.png")
        for i in range(1, rounds):
            embed.add_field(name="Round %s" % i,
                            value="%s - %s takes %s damage! (%s HP remaining)\n%s - %s takes %s damage! (%s HP remaining)"
                                  % (
                                  hitType[i][0], mobInst.info[2], DMG[i][0], HP[i][0], hitType[i][1], intruder_dict[ctx.guild.id][1].info[2],
                                  DMG[i][1], HP[i][1]), inline=False)

        await ctx.send(embed=embed, files=[File, arenaFile, winnerFile])

        if winner == mobInst:
            addCoins = random.randint(0, 11)
            # gain xp

            itemSheet.cell(row=user.row, column=2).value += addCoins
            embed.add_field(name="Rewards", value="%s's %s has gained %s experience points!\n%s has gained %s coins" % (
            mobInst.info[1], mobInst.info[2], xp, mobInst.info[1], addCoins))

            mobSheet.cell(row=mobInst.row, column=7).value += xp
            mobInst.info[6] += xp
            if mobSheet.cell(row=mobInst.row, column=7).value >= 100:
                oldStats = mobInst.stats[:]
                oldInfo = mobInst.info[:]
                mobInst.level_up(1)

                for column in range(8, 13):
                    mobSheet.cell(row=mobInst.row, column=column).value = mobInst.stats[column - 8]

                mobSheet.cell(row=mobInst.row, column=6).value = mobInst.info[5]
                mobSheet.cell(row=mobInst.row, column=7).value = mobInst.info[6]
                newStats = mobInst.stats
                newInfo = mobInst.info
                await print_level_up(ctx.message, oldStats, oldInfo, newStats, newInfo)

        intruder_dict[ctx.guild.id] = [False, None, None]

    else:
        File = discord.File("./Images/Helping Hand.png", filename="image.png")
        embed = discord.Embed(title="There are currently no intruders in your dungeon, Lord %s" % user.name,
                              description="I'm sure one of those wretches will show up eventually", color=0x114a1d)
        embed.set_author(name="Helping Hand", icon_url="attachment://image.png")
        await ctx.send(embed=embed, file=File)
    workbook.save(filename=datafile)


@client.command()
async def sql(ctx):
    print("1")
    db = mysql.connect(
        host="_SOME SQL IMPLEMENTING",
        database="_SOME SQL IMPLEMENTING",
        user="_SOME SQL IMPLEMENTING",
        passwd="_SOME SQL IMPLEMENTING",
        port=SOME SQL IMPLEMENTING
    )
    print(db)


# @client.command()
# async def duel(ctx, member : discord.Member):
# emojis = ['❌','1️⃣','2️⃣','3️⃣','4️⃣','5️⃣','6️⃣','7️⃣', '8️⃣', '9️⃣', '🔟', '✅']
# embed = discord.Embed(title = "A BATTLE OF THE AGES", description = ctx.author.mention + ' V.S ' + member.mention, color = 0x53434)
# embed.set_author(name="Mob Duel")
# embed.add_field(name=ctx.author.name, value="\u200b", inline=True)
# embed.add_field(name=member.name, value="\u200b", inline=True)
# file = discord.File("./Images/duel.png", filename="image.png")
# embed.set_thumbnail(url="attachment://image.png")
# mesg = await ctx.send(file = file, embed=embed)

# for emoji in emojis:
# await mesg.add_reaction(emoji)

# @client.command()
# async def trade(ctx, member:discord.Member):
# emojis = ['❌','1️⃣','2️⃣','3️⃣','4️⃣','5️⃣','6️⃣','7️⃣', '8️⃣', '9️⃣', '🔟', '✅']
# embed = discord.Embed(title = "A Gentleman's Agreement", description = ctx.author.mention + ' trades with ' + member.mention, color = 0x53434)
# embed.set_author(name="Mob Trade")
# embed.add_field(name=ctx.author.name, value='\u200b' , inline=True)
# embed.add_field(name=member.name, value='\u200b', inline=True)
# file = discord.File("./Images/trade.png", filename="image.png")
# embed.set_thumbnail(url="attachment://image.png")
# mesg = await ctx.send(file = file, embed=embed)

# for emoji in emojis:
# await mesg.add_reaction(emoji)

# @client.command()
# async def tunes(ctx):
# channel = ctx.author.voice.channel
# await channel.connect()

client.run('bot_token_here')

